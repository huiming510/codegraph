# -*- coding: utf-8 -*-
"""
# @Time    : 2025/11/5 17:32
# @Author  : cuils
# @Description: Excel 文件解析
"""
import re
import os
import uuid
import base64
import zipfile
import openpyxl
import pandas as pd
import lxml.etree as ET
from typing import List
from pathlib import Path
from unicodedata import normalize
from .base_parser import BaseParser
from ..base import LinkDocument
from ..format_convert import BaseFormatConverter, LibreofficeFormatConverter


class XLSXParser:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        self.archive = zipfile.ZipFile(self.xlsx_path, 'r')
        self.sheet_mapping = self._build_sheet_mapping_from_xml()
        self.sheets = []

    def parse_all(self):
        """解析所有工作表"""
        for ws in self.wb.worksheets:
            self.sheets.append(self.parse_sheet(ws))

    def parse_sheet(self, ws):
        """解析单个工作表的文本、图片、超链接、文本框和形状"""
        table = self._extract_table_and_hyperlink(ws)
        images = self._extract_images(ws)
        shapes = self._extract_shape_and_textboxes(ws)
        sheet_info = {
            "name": ws.title,
            "table": table,
            "images": images,
            "shapes": shapes
        }
        return sheet_info

    def _extract_table_and_hyperlink(self, ws) -> list[list[str]]:
        """提取工作表中的数据及超链接"""
        max_row = ws.max_row
        max_col = ws.max_column
        table = [] # 2维表格
        for rid in range(1, max_row + 1):
            row_values = []
            for cid in range(1, max_col + 1):
                cell = ws.cell(row=rid, column=cid) # start 1
                value = cell.value
                if cell.hyperlink and value is not None:
                    display = str(value)
                    link = cell.hyperlink.target
                    value = f"[{display}]({link})]"
                elif value is None:
                    value = ""
                else:
                    value = str(value)
                row_values.append(value)
            table.append(row_values)

        return table

    def _extract_images(self, ws) -> list[dict[str, str]]:
        images = []
        for _id, _image in enumerate(getattr(ws, "_images"), start=1):
            _byte = _image._data()
            _ext = _image.format.lower()
            _str = base64.b64encode(_byte).decode()

            anchor = _image.anchor
            if hasattr(anchor, "_from"): # start 0
                _row = anchor._from.row + 1
                _col = anchor._from.col + 1
            else:
                _row = ws.max_row + 1
                _col = 1

            images.append(
                {
                    "img_name": f"image-{_id}.{_ext}",
                    "img_ext": _ext,
                    "img_str": _str,
                    "img_loc": [_row, _col]
                }
            )
        return images

    def _extract_shape_and_textboxes(self, ws):
        """抽取形状和文本框中的文本， 这一部分都存储在 xl/drawings/drawing*.xml中"""
        shapes = []
        sheet_xml = self.sheet_mapping.get(ws.title)
        if not sheet_xml:
            return shapes
        rel_xml = f"{os.path.dirname(sheet_xml)}/_rels/{os.path.basename(sheet_xml)}.rels"
        if rel_xml not in self.archive.namelist():
            return shapes

        # 找到该sheet对应的drawing xml
        try:
            drawing_target = None
            rel_root = ET.fromstring(self.archive.read(rel_xml))
            for element in rel_root.iter():
                ele_type = element.get("Type", "")
                if ele_type.endswith("/drawing"):
                    drawing_target = element.get("Target")
                    break

            if not drawing_target:
                return shapes
            drawing_path = drawing_target.lstrip("/").lstrip("../")
            if not drawing_path.startswith("xl"):
                drawing_path = f"xl/{drawing_path}"

            if drawing_path not in self.archive.namelist():
                return shapes
            drawing_xml = self.archive.read(drawing_path)
            shapes.extend(self._parse_drawing_for_text(drawing_xml))
        except Exception as e:
            print(repr(e))
        return shapes

    def _build_sheet_mapping_from_xml(self):
        """解析workbook.xml，建立sheet到XML路径的映射"""
        rel_mapping = {}
        sheet_mapping = {}

        ns = {
            "d": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        # 读取 workbook 关系，获取 rId -> target
        try:
            rel_xml = self.archive.read("xl/_rels/workbook.xml.rels")
            rel_root = ET.fromstring(rel_xml)
            for element in rel_root.iter():
                tag = element.tag
                if tag.endswith("Relationship") or tag == "Relationship":
                    rel_type = element.get("Type")
                    if rel_type and rel_type.endswith("/worksheet"):
                        rel_id = element.get("Id")
                        rel_target = element.get("Target")
                        if rel_id and rel_target:
                            rel_mapping[rel_id] = rel_target
        except Exception as e:
            print(repr(e))
            return sheet_mapping

        # 读取 workbook.xml 获取 sheet 列表
        try:
            wb_xml = self.archive.read("xl/workbook.xml")
            wb_root = ET.fromstring(wb_xml)
            for element in wb_root.iter():
                tag = element.tag

                if tag.endswith("}sheet") or tag == "sheet":
                    name = element.get("name")
                    if not name:
                        continue
                    rel_id = element.get(f"{{{ns['r']}}}id")
                    if not rel_id:
                        rel_id = element.get("r:id")
                    if not rel_id:
                        for attr_name, attr_value in element.attrib.items():
                            if attr_name.endswith("}id") or attr_name == "id":
                                rel_id = attr_value
                                break
                    if rel_id and rel_id in rel_mapping:
                        target = rel_mapping[rel_id].lstrip("/")
                        if not target.startswith(f"xl/"):
                            target = f"xl/{target}"
                        sheet_mapping[name] = target

        except Exception as e:
            print(repr(e))
            return sheet_mapping

        return sheet_mapping

    def _parse_drawing_for_text(self, drawing_xml):
        """从 drawing xml中提取textboxes / shapes 的文字、线条、箭头和锚点"""
        shapes = []

        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        }
        drawing_root = ET.fromstring(drawing_xml)

        anchors = (drawing_root.findall(".//xdr:twoCellAnchor", ns) +
                   drawing_root.findall(".//xdr:oneCellAnchor", ns) +
                   drawing_root.findall(".//xdr:absoluteAnchor", ns))

        for anchor in anchors:
            # 提取文本内容
            texts = []
            for t in anchor.findall(".//a:t", ns):
                if t.text:
                    texts.append(t.text)

            # 提取形状信息（包括线条和箭头）
            shape_info = self._extract_shape_info(anchor, ns)

            # 计算起始锚点位置
            from_node = anchor.find("xdr:from", ns)
            from_cell = None
            if from_node is not None:
                try:
                    col = from_node.findtext("xdr:col", default="0", namespaces=ns) # start 0
                    row = from_node.findtext("xdr:row", default="0", namespaces=ns) # start 0
                    col = int(col) + 1
                    row = int(row) + 1
                    from_cell = [row, col]
                except Exception as e:
                    print(repr(e))
                    from_cell = None

            # 提取终止锚点位置
            to_node = anchor.find("xdr:to", ns)
            to_cell = None
            if to_node is not None:
                try:
                    col = to_node.findtext("xdr:col", default="0", namespaces=ns) # start 0
                    row = to_node.findtext("xdr:row", default="0", namespaces=ns) # start 0
                    col = int(col) + 1
                    row = int(row) + 1
                    to_cell = [row, col]
                except Exception as e:
                    print(repr(e))
                    to_cell = None

            if texts:
                shape = {
                    "type": "textbox",
                    "text": "".join(texts).strip(),
                    "from_cell": from_cell,
                    "to_cell": to_cell,
                    **shape_info
                }
                shapes.append(shape)
            elif shape_info.get("shape_type") in ["line", "arrow"]:
                shape = {
                    "type": shape_info.get("shape_type"),
                    "text": "",
                    "from_cell": from_cell,
                    "to_cell": to_cell,
                    **shape_info
                }
                shapes.append(shape)
        return shapes

    def _extract_shape_info(self, anchor, ns):
        """抽取形状信息，包括线条和箭头类型"""
        shape_info = {
            "shape_type": None,
            "prst": None
        }
        # 查找 shape 元素
        sp = anchor.find("xdr:sp", ns)
        if sp is None:
            return shape_info

        # 查找形状属性
        sp_pr = sp.find("xdr:spPr", ns)
        if sp_pr is None:
            return shape_info

        # 查找预设几何形状
        prst_geom = sp_pr.find("a:prstGeom", ns)
        if prst_geom is not None:
            prst = prst_geom.get("prst")
            if prst:
                shape_info["prst"] = prst
                # 判断是否为线条或箭头
                prst_lower = prst.lower()
                if "line" in prst_lower and "arrow" not in prst_lower:
                    shape_info["shape_type"] = "line"
                elif "arrow" in prst_lower or prst_lower in ["rightArrow", "leftArrow", "upArrow", "downArrow",
                                                             "bentArrow", "uturnArrow", "curvedRightArrow",
                                                             "curvedLeftArrow", "curvedUpArrow", "curvedDownArrow",
                                                             "stripedRightArrow", "notchedRightArrow"]:
                    shape_info["shape_type"] = "arrow"
                else:
                    shape_info["shape_type"] = "shape"

        # 查找自定义几何形状（可能包含线条路径）
        cust_geom = sp_pr.find("a:custGeom", ns)
        if cust_geom is not None:
            # 检查路径列表，如果只有简单的线条路径，可能是线条
            path_list = cust_geom.find("a:pathLst", ns)
            if path_list is not None:
                paths = path_list.findall("a:path", ns)
                # 如果路径简单（只有 moveTo 和 lnTo），可能是线条
                if len(paths) == 1:
                    path = paths[0]
                    move_tos = path.findall("a:moveTo", ns)
                    ln_tos = path.findall("a:lnTo", ns)
                    close = path.find("a:close", ns)
                    # 如果有 moveTo 和 lnTo，但没有 close，可能是线条
                    if move_tos and ln_tos and close is None:
                        shape_info["shape_type"] = "line"
                        shape_info["prst"] = "customLine"

        return shape_info

    def _convert_sheet_to_markdown(self, sheet_info):
        """将sheet转为markdown"""
        table = sheet_info["table"]
        images = sheet_info["images"]
        shapes = sheet_info["shapes"]
        max_row, max_col = len(table), len(table[0])
        extra = ""
        for image in images:
            row, col = image["img_loc"]
            if 1<=row<=max_row and 1<=col<=max_col:
                table[row-1][col-1] += f"\n![media/image]({image['img_name']})\n"
            else:
                extra += f"\n![media/image]({image['img_name']})\n"

        for shape in shapes:
            # 暂时只考虑文本框，箭头还不知道怎么用
            if shape["type"] == "textbox":
                text = shape["text"]
                row, col = shape["from_cell"]
                if 1 <= row <= max_row and 1 <= col <= max_col:
                    table[row-1][col-1] += f"\n**[textbox]** {text}\n"
                else:
                    extra += f"\n**[textbox]** {text}\n"

        # 删除多余的空行，和空列
        filter_row_table = []
        for row_cells in table:
            row_cells = [value.strip() for value in row_cells]
            # 删除掉全为空的行
            if all(not value for value in row_cells):
                continue
            filter_row_table.append(row_cells)
        filter_col_table = []
        min_col, max_col = 1, max_col
        while min_col < max_col:
            if all(not filter_row_table[i][min_col-1] for i in range(len(filter_row_table))):
                min_col += 1
            else:
                break
        while min_col < max_col:
            if all(not filter_row_table[i][max_col-1] for i in range(len(filter_row_table))):
                max_col -= 1
            else:
                break
        for row_cells in filter_row_table:
            filter_col_table.append(row_cells[min_col-1:max_col])

        header = filter_col_table[0]
        lines = [f"|{'|'.join(header)}|"]
        separator = "|".join(["---" for _ in header])
        lines.append(f"|{separator}|")

        for row in filter_col_table[1:]:
            lines.append(f"|{'|'.join(row)}|")

        markdown = f"## {sheet_info['name']}\n\n"+"\n".join(lines)+extra
        return markdown

    def convert_to_markdown(self):
        for sheet_info in self.sheets:
            markdown = self._convert_sheet_to_markdown(sheet_info)
            sheet_info["markdown"] = markdown
        return self.sheets


class ExcelParser(BaseParser):
    def __init__(self, logger, converter: BaseFormatConverter=None):
        super().__init__(logger)
        self.ignore_sheet_name_keywords = ["Evaluation Warning"]  # 忽略 sheet
        self.converter = converter
        if self.converter is None:
            self.converter = LibreofficeFormatConverter()

    @property
    def support_file_formats(self) -> List[str]:
        return ["xls", "xlsx", "xlsm", "csv"]

    def parse(self, filepath: str | Path, doc_id:str=None) -> List[LinkDocument]:
        filepath = Path(filepath).absolute()
        suffix = filepath.suffix.lstrip(".")
        if suffix not in self.support_file_formats:
            raise ValueError(f"Excel Parser only supports {self.support_file_formats}, but got {suffix}")

        # 后缀为xls，且不支持文件格式转换，则使用兜底方式
        new_filepath = None
        if suffix == "xls":
            try:
                xlsx_file = self.converter.convert(
                    input_filepath=filepath,
                    input_format="xls",
                    output_format="xlsx",
                    output_suffix="xlsx"
                )
                new_filepath = Path(xlsx_file).absolute()
            except Exception as e:
                self.logger.info(f"[Warning]: 转换失败，尝试使用pandas解析。`{filepath}` - {repr(e)}")
                return self._parse_by_pandas(filepath)

        try:
            return self._parse_by_openpyxl(filepath, new_filepath=new_filepath, doc_id=doc_id)
        except Exception as e:
            self.logger.info(f"[Warning]: openpyxl解析失败，尝试使用pandas解析。`{filepath}` - {repr(e)}")
            return self._parse_by_pandas(filepath, doc_id=doc_id)

    def _parse_by_pandas(self, filepath: Path, doc_id=None) -> List[LinkDocument]:
        """pandas 解析 xls和csv文件，以及xlsx的兜底方案"""
        try:
            df_book = pd.read_excel(filepath, sheet_name=None, header=None)
        except Exception as e:
            self.logger.info(f"[Warning]: pandas读取excel失败，尝试读取csv。{filepath}` - {repr(e)}")
            try:
                sheet = pd.read_csv(filepath)
                df_book = {"csv_only_first_sheet": sheet}
            except Exception as ee:
                self.logger.info(f"[Read Error]: 文件读取失败。`{filepath}` - {repr(ee)}")
                return []

        documents = []
        for i, sheet_name in enumerate(df_book, start=1):
            is_skip = False
            for keyword in self.ignore_sheet_name_keywords:
                if keyword in sheet_name:
                    is_skip = True
                    break
            if is_skip:
                continue

            df_sheet = df_book[sheet_name]
            content = self._parse_sheet_by_pandas(df_sheet)

            document = LinkDocument(
                doc_id=f"{doc_id}_sheet_{i}" if doc_id else str(uuid.uuid4()),
                content=content,
                metadata={"file_name": filepath.name, "sheet_name": sheet_name,},
            )
            documents.append(document)

        return documents

    def _parse_sheet_by_pandas(self, df_sheet: pd.DataFrame) -> str:
        """将sheet转换为文本"""
        df_sheet = df_sheet.map(lambda x: x.strip().replace("\n", "") if isinstance(x, str) else x, na_action="ignore")
        content = df_sheet.to_csv(index=False, header=False, na_rep="", sep="\t")

        lines = []
        for line in content.splitlines():
            # 删除右侧多余sep
            line = line.rstrip()
            # 空行
            if not line.strip():
                if not lines:
                    continue
                lines.append("")

            # 将左侧sep替换成空格
            replace_space_end_index = 0
            for i, c in enumerate(line):
                if c != "\t":
                    replace_space_end_index = i
                    break
            line = " " * replace_space_end_index + line[replace_space_end_index:]
            lines.append(line)

        content = "\n".join(lines)
        # 将中间sep全部替换成一个\t
        content = re.sub(r"\t{2,}", "\t\t", content)
        content = re.sub(r"\n{2,}", "\n\n", content)
        return normalize("NFKC", content)

    def _parse_by_openpyxl(self, filepath: Path, new_filepath: Path = None, doc_id=None) -> List[LinkDocument]:
        """openpyxl解析"""
        if new_filepath:
            # 该参数说明，是xls转xlsx后的文件路径
            xlsx_file = XLSXParser(xlsx_path=new_filepath)
        else:
            xlsx_file = XLSXParser(filepath)

        xlsx_file.parse_all()

        documents = []
        for i, sheet_info in enumerate(xlsx_file.convert_to_markdown(), start=1):
            document = LinkDocument(
                doc_id=doc_id,
                content=sheet_info["markdown"],
                images=sheet_info["images"],
                metadata={"file_name": filepath.name, "sheet_name": sheet_info["name"], "sheet_idx": i},
            )
            documents.append(document)

        return documents